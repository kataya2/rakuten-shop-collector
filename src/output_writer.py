from __future__ import annotations
import csv
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from src.utils import get_logger

if TYPE_CHECKING:
    from src.shop_extractor import ShopInfo

logger = get_logger(__name__)

HEADERS = [
    "shop_id", "shop_name", "shop_url", "item_count",
    "avg_review", "total_reviews", "min_price", "genre_id",
]


def write_csv(shops: list[ShopInfo], filepath: str) -> None:
    """ショップリストをUTF-8 BOM付きCSVに書き出す（Excel対応）。"""
    Path(filepath).parent.mkdir(parents=True, exist_ok=True)
    with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=HEADERS)
        writer.writeheader()
        for s in shops:
            writer.writerow({
                "shop_id": s.shop_id,
                "shop_name": s.shop_name,
                "shop_url": s.shop_url,
                "item_count": s.item_count,
                "avg_review": s.avg_review,
                "total_reviews": s.total_reviews,
                "min_price": s.min_price,
                "genre_id": s.genre_id,
            })
    logger.info("CSV出力完了: %s", filepath)


def write_excel(shops: list[ShopInfo], filepath: str) -> None:
    """ショップリストを書式付きExcelファイルに書き出す。"""
    Path(filepath).parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "shops"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center")

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for s in shops:
        ws.append([
            s.shop_id, s.shop_name, s.shop_url, s.item_count,
            s.avg_review, s.total_reviews, s.min_price, s.genre_id,
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    wb.save(filepath)
    logger.info("Excel出力完了: %s", filepath)


def write_gsheet(shops: list[ShopInfo], sheet_id: str, credentials_path: str) -> None:
    """ショップリストをGoogle Spreadsheetsの'shops'シートに上書き出力する。"""
    import gspread
    from google.oauth2.service_account import Credentials

    scopes = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(credentials_path, scopes=scopes)
    gc = gspread.authorize(creds)
    spreadsheet = gc.open_by_key(sheet_id)

    try:
        ws = spreadsheet.worksheet("shops")
    except gspread.exceptions.WorksheetNotFound:
        ws = spreadsheet.add_worksheet(title="shops", rows=1000, cols=len(HEADERS))

    ws.clear()
    rows = [HEADERS] + [
        [s.shop_id, s.shop_name, s.shop_url, s.item_count,
         s.avg_review, s.total_reviews, s.min_price, s.genre_id]
        for s in shops
    ]
    ws.update(rows)
    logger.info("Googleスプレッドシート出力完了: %s", sheet_id)
