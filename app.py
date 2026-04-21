"""Streamlit GUI for Rakuten Shop Collector."""
from __future__ import annotations
import csv
import io
import math
import os

import pandas as pd
import streamlit as st
from dotenv import load_dotenv

from src.api_client import ApiClient
from src.output_writer import HEADERS
from src.shop_extractor import ShopInfo, extract_shops
from src.utils import RakutenAPIError

load_dotenv()

st.set_page_config(
    page_title="Rakuten Shop Collector",
    page_icon="🛍️",
    layout="wide",
)


# ---------------------------------------------------------------------------
# Internal helpers (in-memory output generation for download buttons)
# ---------------------------------------------------------------------------

class _ProgressApiClient(ApiClient):
    """ApiClientのサブクラス: ページ取得ごとにコールバックを呼ぶ。"""

    def __init__(self, *args, on_page=None, **kwargs):
        super().__init__(*args, **kwargs)
        self._on_page = on_page

    def _get_with_retry(self, params):
        result = super()._get_with_retry(params)
        if self._on_page:
            self._on_page()
        return result


def _to_dataframe(shops: list[ShopInfo]) -> pd.DataFrame:
    return pd.DataFrame([
        {
            "shop_id": s.shop_id,
            "shop_name": s.shop_name,
            "shop_url": s.shop_url,
            "item_count": s.item_count,
            "avg_review": s.avg_review,
            "total_reviews": s.total_reviews,
            "min_price": s.min_price,
            "genre_id": s.genre_id,
        }
        for s in shops
    ])


def _to_csv_bytes(shops: list[ShopInfo]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=HEADERS)
    writer.writeheader()
    for s in shops:
        writer.writerow({
            "shop_id": s.shop_id, "shop_name": s.shop_name,
            "shop_url": s.shop_url, "item_count": s.item_count,
            "avg_review": s.avg_review, "total_reviews": s.total_reviews,
            "min_price": s.min_price, "genre_id": s.genre_id,
        })
    return buf.getvalue().encode("utf-8-sig")


def _to_excel_bytes(shops: list[ShopInfo]) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "shops"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")
    for s in shops:
        ws.append([s.shop_id, s.shop_name, s.shop_url, s.item_count,
                   s.avg_review, s.total_reviews, s.min_price, s.genre_id])
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------

st.title("🛍️ Rakuten Shop Collector")
st.caption("楽天市場からショップ情報を収集し、CSV / Excel 形式でダウンロードできます。")

with st.sidebar:
    st.header("⚙️ 検索設定")

    search_type = st.radio("検索方法", ["キーワード", "カテゴリID"], horizontal=True)

    if search_type == "キーワード":
        keyword: str | None = st.text_input("キーワード", placeholder="例: ワイヤレスイヤホン") or None
        category_id: str | None = None
    else:
        keyword = None
        category_id = st.text_input("カテゴリID", placeholder="例: 100371") or None

    count = st.slider("取得件数", min_value=30, max_value=500, value=30, step=10)

    st.divider()
    st.caption("認証情報は `.env` から自動読み込みされます。")

run_btn = st.button("🔍 検索実行", type="primary", use_container_width=True)

if run_btn:
    if not keyword and not category_id:
        st.error("キーワードまたはカテゴリIDを入力してください。")
        st.stop()

    app_id = os.environ.get("RAKUTEN_APP_ID", "")
    access_key = os.environ.get("RAKUTEN_ACCESS_KEY", "")
    if not app_id or not access_key:
        st.error(
            "認証情報が設定されていません。"
            "`.env` に `RAKUTEN_APP_ID` と `RAKUTEN_ACCESS_KEY` を設定してください。"
        )
        st.stop()

    referer = os.environ.get("RAKUTEN_REFERER", "https://github.com")
    max_pages = math.ceil(count / 30)
    progress_bar = st.progress(0, text="検索中...")
    fetched = [0]

    def _on_page() -> None:
        fetched[0] += 1
        pct = min(fetched[0] / max_pages, 1.0)
        progress_bar.progress(pct, text=f"取得中... {fetched[0]}/{max_pages} ページ")

    try:
        client = _ProgressApiClient(
            app_id, access_key, referer=referer, on_page=_on_page
        )
        items = client.search(keyword=keyword, category_id=category_id, count=count)
        shops = extract_shops(items)
        progress_bar.progress(1.0, text="完了")
    except RakutenAPIError as e:
        progress_bar.empty()
        st.error(f"APIエラー: {e}")
        st.stop()
    except Exception as e:
        progress_bar.empty()
        st.error(f"予期しないエラーが発生しました: {e}")
        st.stop()

    st.session_state["shops"] = shops

if "shops" in st.session_state:
    shops: list[ShopInfo] = st.session_state["shops"]

    if not shops:
        st.warning("検索結果が0件でした。")
        st.stop()

    st.success(f"{len(shops)} ショップが見つかりました（商品数の多い順）")

    df = _to_dataframe(shops)
    st.dataframe(df, use_container_width=True, hide_index=True)

    dl_col1, dl_col2 = st.columns(2)
    with dl_col1:
        st.download_button(
            "📥 CSVダウンロード",
            data=_to_csv_bytes(shops),
            file_name="shops_result.csv",
            mime="text/csv",
            use_container_width=True,
        )
    with dl_col2:
        st.download_button(
            "📥 Excelダウンロード",
            data=_to_excel_bytes(shops),
            file_name="shops_result.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    st.subheader("📊 分析グラフ")
    chart_col1, chart_col2 = st.columns(2)

    with chart_col1:
        st.markdown("**平均レビュースコア分布**")
        valid_reviews = df[df["avg_review"] > 0]["avg_review"]
        if len(valid_reviews) > 0:
            bins = [0.0, 1.0, 2.0, 3.0, 3.5, 4.0, 4.5, 5.01]
            labels = ["~1", "~2", "~3", "~3.5", "~4", "~4.5", "~5"]
            hist = pd.cut(valid_reviews, bins=bins, labels=labels, right=True)
            st.bar_chart(hist.value_counts().sort_index().rename("ショップ数"))
        else:
            st.info("レビューデータがありません")

    with chart_col2:
        st.markdown("**商品数上位10ショップ**")
        top10 = (
            df.nlargest(10, "item_count")
            .set_index("shop_name")["item_count"]
            .rename("商品数")
        )
        st.bar_chart(top10)
